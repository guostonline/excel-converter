import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, GradientFill
import json


class Excel:

    def __init__(self, path, rest_days=None):
        self.__day_work = 24
        self.path = path
        self.rest_days = rest_days
       

    def get_day_work(self) -> tuple:
        """
        return tuple as total days of month and day works
        """
        wb = load_workbook(self.path)
        sheet_ranges = wb["AGADIR"]
        # wb.active
        day_work = sheet_ranges["C6"].value
        day_work = day_work.split(" ", 2)
        a: str = day_work[0].replace("/", "")
        b: str = day_work[1]

        print(f"day work is : {int(a), int(b)}")
        with open("days.json", "r") as jsonFile:
            data = json.load(jsonFile)

        data["from_file"] = {"t": b, "d": a}

        with open("days.json", "w") as jsonFile:
            json.dump(data, jsonFile)

        return int(a), int(b.strip())

    def fix_sheet(self, jour_rest=None):
        # Use custom rest_days if available, otherwise use jour_rest parameter
        effective_jour_rest = self.rest_days if self.rest_days is not None else jour_rest
        
        # Load days.json to get the "d" value
        try:
            with open('days.json', 'r') as f:
                days_data = json.load(f)
                d_value = int(days_data["from_file"]["d"])
        except (FileNotFoundError, KeyError, ValueError):
            d_value = 4  # Default fallback value

        wb = load_workbook(self.path)
        sheet_ranges_quali = wb["QUALI NV"]
        sheet_ranges_quali.unmerge_cells("E1:K2")
        sheet_ranges_quali.delete_rows(1, 7)
        sheet_ranges_quali.delete_rows(2, 4)
        sheet_ranges_quali.delete_rows(10, 1)
        sheet_ranges_quali.delete_cols(1, 3)
        sheet_ranges_quali.delete_cols(2, 3)
        sheet_ranges_quali.delete_cols(3, 3)
        sheet_ranges_quali.delete_cols(4, 11)
        sheet_ranges_quali.delete_cols(7, 2)
        sheet_ranges_quali.delete_rows(sheet_ranges_quali.max_row - 1)

        sheet_ranges_quali['A1'] = "Vendeur"
        sheet_ranges_quali['C1'] = "ACM"
        sheet_ranges_quali['F1'] = "LINE"
        sheet_ranges_quali['G1'] = "TSM"
        sheet_ranges_quali['G1'].fill = PatternFill("solid", fgColor="4cbb17")
        sheet_ranges_quali["F1"].fill = PatternFill("solid", fgColor="4cbb17")
        ## AGADIR
        sheet_ranges_quanti = wb["AGADIR"]
        sheet_ranges_quanti.unmerge_cells("A8:A9")
        sheet_ranges_quanti.unmerge_cells("B8:B9")
        sheet_ranges_quanti.unmerge_cells("D8:D9")
        sheet_ranges_quanti.unmerge_cells("F8:J8")
        sheet_ranges_quanti.unmerge_cells("K8:O8")
        #delete columns
        sheet_ranges_quanti.delete_cols(1, 2) 
        sheet_ranges_quanti.delete_cols(3, 1) 
        sheet_ranges_quanti.delete_cols(6, 2) 
        sheet_ranges_quanti.delete_cols(7, 2) 
        sheet_ranges_quanti.delete_cols(9, 1)  
        sheet_ranges_quanti.delete_cols(10, 1) 
        #delete rowsre
        sheet_ranges_quanti.delete_rows(1, 8)
        sheet_ranges_quanti.delete_rows(2, 32)
        sheet_ranges_quanti.delete_rows(154, 8)
        sheet_ranges_quanti.delete_rows(170, 14)
        sheet_ranges_quanti['A1'] = "Vendeur"
        sheet_ranges_quanti['B1'] = "Famille"
        sheet_ranges_quanti['C1'] = "REAL"
        sheet_ranges_quanti['D1'] = "OBJ"
        sheet_ranges_quanti['E1'] = "Percent"
        sheet_ranges_quanti['F1'] = "REAL 2025"
        sheet_ranges_quanti['G1'] = "H 2024"
        sheet_ranges_quanti['H1'] = "H %"
        sheet_ranges_quanti['I1'] = "EnCours"
        sheet_ranges_quanti['J1'] = "OBJ MOIS"
        sheet_ranges_quanti['K1'] = "RAF"
        
        

        for i in range(sheet_ranges_quanti.max_row):
            # Replace '%' strings with None for proper numeric handling
            if sheet_ranges_quanti[f"E{i + 1}"].value == '%':
                sheet_ranges_quanti[f"E{i + 1}"].value = None
            # Convert percentage values to decimal (e.g., 85% -> 0.85)
            elif isinstance(sheet_ranges_quanti[f"E{i + 1}"].value, (int, float)) and i > 0:
                # If value is greater than 1, assume it's a percentage and convert to decimal
                if sheet_ranges_quanti[f"E{i + 1}"].value > 1:
                    sheet_ranges_quanti[f"E{i + 1}"].value = sheet_ranges_quanti[f"E{i + 1}"].value / 100
            
            if sheet_ranges_quanti[f"B{i + 1}"].value == 'SAUCES TACOS':
                sheet_ranges_quanti[f"B{i + 1}"].value = 'SAUCES'

            # Add EnCours values to REAL column (C = REAL + EnCours)
            if i > 0:  # Skip header row
                real_cell = sheet_ranges_quanti[f"C{i + 1}"].value
                encours_cell = sheet_ranges_quanti[f"I{i + 1}"].value
                
                # Add EnCours to REAL if both are numeric
                if isinstance(real_cell, (int, float)) and isinstance(encours_cell, (int, float)):
                    sheet_ranges_quanti[f"C{i + 1}"].value = real_cell + encours_cell
                elif isinstance(encours_cell, (int, float)) and (real_cell is None or real_cell == 0):
                    # If REAL is empty/zero but EnCours has value, use EnCours
                    sheet_ranges_quanti[f"C{i + 1}"].value = encours_cell

            # Calculate OBJ MOIS = (OBJ * self.__day_work) / d_value
            if sheet_ranges_quanti[f"J{i + 1}"].value == '%':
                sheet_ranges_quanti[f"J{i + 1}"].value = None
                obj_cell = sheet_ranges_quanti[f"D{i + 1}"].value
                if isinstance(obj_cell, (int, float)):
                    # OBJ MOIS = (OBJ * self.__day_work) / d_value
                    sheet_ranges_quanti[f"J{i + 1}"].value = (obj_cell * self.__day_work) / d_value
                else:
                    sheet_ranges_quanti[f"J{i + 1}"].value = None

            # Calculate RAF = (OBJ - REAL) / rest_days
            if sheet_ranges_quanti[f"K{i + 1}"].value == '%':
                sheet_ranges_quanti[f"K{i + 1}"].value = None
                obj_cell = sheet_ranges_quanti[f"D{i + 1}"].value
                real_cell = sheet_ranges_quanti[f"C{i + 1}"].value
                if isinstance(obj_cell, (int, float)) and isinstance(real_cell, (int, float)):
                    # RAF = (OBJ - REAL) / rest_days
                    rest_days = effective_jour_rest if effective_jour_rest is not None else d_value
                    sheet_ranges_quanti[f"K{i + 1}"].value = (obj_cell - real_cell) / rest_days
                else:
                    sheet_ranges_quanti[f"K{i + 1}"].value = None

        wb.save("excel/finale_jour.xlsx")
        return True
        