import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, GradientFill
import json
import pandas as pd
import os
from google_sheets import GoogleSheetsService


class Excel:

    def __init__(self, path, rest_days=None):
        self.__day_work = 24
        self.path = path
        self.rest_days = rest_days
        self.ttc_rate = 1,2

    def get_day_work(self) -> tuple:
        """
        return tuple as total days of month and day works
        """
        wb = load_workbook(self.path)
        sheet_ranges = wb["AGADIR"]
        # wb.active
        day_work = sheet_ranges["C6"].value
        day_work = day_work.split(" ", 2)
        a: str = day_work[0].replace("/", "")
        b: str = day_work[1]

        print(f"day work is : {int(a), int(b)}")
        with open("days.json", "r") as jsonFile:
            data = json.load(jsonFile)

        data["from_file"] = {"t": b, "d": a}

        with open("days.json", "w") as jsonFile:
            json.dump(data, jsonFile)

        return int(a), int(b.strip())

    

    def fix_sheet(self, jour_rest=None):
        # Use custom rest_days if available, otherwise use jour_rest parameter
        

        wb = load_workbook(self.path)
        sheet_ranges_quali = wb["QUALI NV"]
        sheet_ranges_quali.unmerge_cells("E1:K2")
        sheet_ranges_quali.delete_rows(1, 7)
        sheet_ranges_quali.delete_rows(2, 4)
        sheet_ranges_quali.delete_rows(10, 1)
        sheet_ranges_quali.delete_cols(1, 3)
        sheet_ranges_quali.delete_cols(2, 3)
        sheet_ranges_quali.delete_cols(3, 3)
        sheet_ranges_quali.delete_cols(4, 11)
        sheet_ranges_quali.delete_cols(7, 2)
        sheet_ranges_quali.delete_rows(sheet_ranges_quali.max_row - 1)

        sheet_ranges_quali['A1'] = "Vendeur"
        sheet_ranges_quali['C1'] = "ACM"
        sheet_ranges_quali['F1'] = "LINE"
        sheet_ranges_quali['G1'] = "TSM"
        sheet_ranges_quali['H1'] = "RAF TSM"
        sheet_ranges_quali['I1'] = "RAF ACM"
        sheet_ranges_quanti['K21'] = "CHAKIB ELFIL"
        sheet_ranges_quanti['K10'] = "BOUTMEZGUINE EL MOSTAFA"
        # Calculate RAF TSM using cell values instead of cell objects
       
            # Compute RAF TSM = (TSM - TSM * H%) / rest_days
        for row in range(2, sheet_ranges_quali.max_row + 1):
            tsm_val = sheet_ranges_quali[f"G{row}"].value
            client_number = sheet_ranges_quali[f"B{row}"].value
            if isinstance(tsm_val, (int, float)) and isinstance(client_number, (int, float)) and jour_rest:
                sheet_ranges_quali[f"H{row}"].value = int((client_number - (client_number * tsm_val)) / jour_rest)

            else:
                sheet_ranges_quali[f"H{row}"].value = None
             # Compute RAF ACM = (ACM - ACM * H%) / rest_days    
        for row in range(2, sheet_ranges_quali.max_row + 1):
            acm_val = sheet_ranges_quali[f"C{row}"].value
            client_number = sheet_ranges_quali[f"B{row}"].value
            if isinstance(acm_val, (int, float)) and isinstance(client_number, (int, float)) and jour_rest:
                sheet_ranges_quali[f"I{row}"].value = int((client_number - (client_number * acm_val)) / jour_rest)

            else:
                sheet_ranges_quali[f"I{row}"].value = None
        sheet_ranges_quali['G1'].fill = PatternFill("solid", fgColor="4cbb17")
        sheet_ranges_quali["F1"].fill = PatternFill("solid", fgColor="4cbb17")
        ## AGADIR
        sheet_ranges_quanti = wb["AGADIR"]
        sheet_ranges_quanti.unmerge_cells("A8:A9")
        sheet_ranges_quanti.unmerge_cells("B8:B9")
        sheet_ranges_quanti.unmerge_cells("D8:D9")
        sheet_ranges_quanti.unmerge_cells("F8:J8")
        sheet_ranges_quanti.unmerge_cells("K8:O8")
        #delete columns
        sheet_ranges_quanti.delete_cols(1, 2) 
        sheet_ranges_quanti.delete_cols(3, 1) 
        sheet_ranges_quanti.delete_cols(6, 2) 
        sheet_ranges_quanti.delete_cols(7, 2) 
        sheet_ranges_quanti.delete_cols(9, 1)  
        sheet_ranges_quanti.delete_cols(10, 1) 
        #delete rowsre
        sheet_ranges_quanti.delete_rows(1, 8)
        sheet_ranges_quanti.delete_rows(2, 32)
        sheet_ranges_quanti.delete_rows(154, 8)
        sheet_ranges_quanti.delete_rows(170, 14)
        sheet_ranges_quanti['A1'] = "Vendeur"
        sheet_ranges_quanti['B1'] = "Famille"
        sheet_ranges_quanti['C1'] = "REAL"
        sheet_ranges_quanti['D1'] = "OBJ"
        sheet_ranges_quanti['E1'] = "Percent"
        sheet_ranges_quanti['F1'] = "REAL 2025"

        sheet_ranges_quanti['G1'] = "H 2024"
        sheet_ranges_quanti['H1'] = "H %"
        sheet_ranges_quanti['I1'] = "EnCours"
        sheet_ranges_quanti['J1'] = "OBJ MOIS"
        sheet_ranges_quanti['K1'] = "RAF"
        sheet_ranges_quanti['K21'] = "CHAKIB ELFIL"
        sheet_ranges_quanti['K10'] = "BOUTMEZGUINE EL MOSTAFA"
        

        for i in range(sheet_ranges_quanti.max_row):
            # Replace '%' strings with None for proper numeric handling
            if sheet_ranges_quanti[f"E{i + 1}"].value == '%':
                sheet_ranges_quanti[f"E{i + 1}"].value = None
            # Convert percentage values to decimal (e.g., 85% -> 0.85)
            elif isinstance(sheet_ranges_quanti[f"E{i + 1}"].value, (int, float)) and i > 0:
                # If value is greater than 1, assume it's a percentage and convert to decimal
                if sheet_ranges_quanti[f"E{i + 1}"].value > 1:
                    sheet_ranges_quanti[f"E{i + 1}"].value = sheet_ranges_quanti[f"E{i + 1}"].value / 100
            
            if sheet_ranges_quanti[f"B{i + 1}"].value == 'SAUCES TACOS':
                sheet_ranges_quanti[f"B{i + 1}"].value = 'SAUCES'

            # Collect all values from column I into an array
            # Add all values from column I to corresponding cells in column C
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                c_val = sheet_ranges_quanti[f"C{row}"].value
                i_val = sheet_ranges_quanti[f"I{row}"].value
                if isinstance(c_val, (int, float)) and isinstance(i_val, (int, float)):
                    sheet_ranges_quanti[f"C{row}"].value = c_val + i_val
            
            
            
            # Compute Percent column: (REAL / OBJ) - 1
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                real_val = sheet_ranges_quanti[f"C{row}"].value
                obj_val = sheet_ranges_quanti[f"D{row}"].value
                if isinstance(real_val, (int, float)) and isinstance(obj_val, (int, float)) and obj_val != 0:
                    sheet_ranges_quanti[f"E{row}"].value = (real_val / obj_val) - 1
                else:
                    sheet_ranges_quanti[f"E{row}"].value = None
            
            # Load the days.json file to get the "b" value
            with open("days.json", "r") as jsonFile:
                data = json.load(jsonFile)
            b_value = data["from_file"]["d"]  # "d" holds the "b" value (total days in month)

            # Compute OBJ MOIS = OBJ * self.__day_work / b
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                obj_val = sheet_ranges_quanti[f"D{row}"].value
                if isinstance(obj_val, (int, float)):
                    sheet_ranges_quanti[f"J{row}"].value = obj_val * self.__day_work / int(b_value)
                else:
                    sheet_ranges_quanti[f"J{row}"].value = None


            # Compute RAF = (OBJ MOIS - REAL) / rest_days
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                obj_mois_val = sheet_ranges_quanti[f"J{row}"].value
                real_val = sheet_ranges_quanti[f"C{row}"].value
                if isinstance(obj_mois_val, (int, float)) and isinstance(real_val, (int, float)) and jour_rest:
                    sheet_ranges_quanti[f"K{row}"].value = (obj_mois_val - real_val) / jour_rest
                else:
                    sheet_ranges_quanti[f"K{row}"].value = None
           
           
            
            # Convert REAL, OBJ, OBJ MOIS, RAF columns to integers
            for row in range(2, sheet_ranges_quanti.max_row + 1):
                for col in ["C", "D", "J", "K","F","G","I" ]:
                    cell = sheet_ranges_quanti[f"{col}{row}"]
                    if isinstance(cell.value, (int, float)):
                        cell.value = int(cell.value)
            wb.save("excel/finale_jour.xlsx")
            return True
    
    def get_quali_nv_dataframe(self):
        """
        Extract QUALI NV sheet data as a DataFrame
        """
        try:
            # Read the QUALI NV sheet from the processed file
            output_path = "excel/finale_jour.xlsx"
            if os.path.exists(output_path):
                df_quali = pd.read_excel(output_path, sheet_name='QUALI NV')
                return df_quali
            else:
                # If processed file doesn't exist, read from original
                df_quali = pd.read_excel(self.path, sheet_name='QUALI NV')
                return df_quali
        except Exception as e:
            print(f"Error reading QUALI NV sheet: {e}")
            return None
            
            
        
        
